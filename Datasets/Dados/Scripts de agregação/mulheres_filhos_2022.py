from openpyxl import load_workbook
import pandas as pd

NORDESTE = ['AL', 'BA', 'CE', 'MA', 'PB', 'PE', 'PI', 'RN', 'SE']

wb = load_workbook('Tabela_10_Mulheres_com_filhos_nascidos_vivos.xlsx', read_only=True)
ws = wb['Tabela SIDRA 10077 e 10078']

rows = []
municipios_vistos = set()

for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i < 7:
        continue

    val = row[0]
    if not val or '(' not in str(val):
        continue

    # Pula percentuais — valores float pequenos
    if row[1] is not None and isinstance(row[1], float) and row[1] < 100:
        continue

    uf = str(val).split('(')[-1].replace(')', '').strip()
    if uf not in NORDESTE:
        continue

    nome = str(val).split('(')[0].strip()
    chave = (nome, uf)

    # Pega só a primeira ocorrência de cada município
    if chave in municipios_vistos:
        continue
    municipios_vistos.add(chave)

    # Seção Total: colunas 1 a 13
    filhos_12_14 = row[1] or 0
    filhos_15_19 = row[2] or 0
    demais       = sum(row[k] or 0 for k in range(3, 14))

    rows.append({
        'NO_MUNICIPIO_IBGE':     nome,
        'SG_UF':                 uf,
        'MULHERES_FILHOS_12_19': filhos_12_14 + filhos_15_19,
        'MULHERES_FILHOS_TOTAL': filhos_12_14 + filhos_15_19 + demais
    })

df_maes = pd.DataFrame(rows)

print(df_maes.shape)
print(df_maes.head())

df_maes.to_csv('maes_jovens_nordeste_2022.csv', sep=';', encoding='utf-8', index=False)