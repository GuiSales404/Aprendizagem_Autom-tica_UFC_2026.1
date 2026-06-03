from openpyxl import load_workbook
import pandas as pd
import os

ESTADOS = ['AL', 'BA', 'CE', 'MA', 'PB', 'PE', 'PI', 'RN', 'SE']

frames = []

for uf in ESTADOS:
    arquivo = f'tabela_MUN_{uf}.xlsx'

    if not os.path.exists(arquivo):
        print(f'{uf}: arquivo não encontrado, pulando...')
        continue

    wb = load_workbook(arquivo, read_only=True)
    ws = wb['Tabela']

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 6:
            continue

        val = row[0]
        if not val or '(' not in str(val):
            continue

        uf_row = str(val).split('(')[-1].replace(')', '').strip()
        nome   = str(val).split('(')[0].strip()

        frames.append({
            'NO_MUNICIPIO_IBGE':       nome,
            'SG_UF':                   uf_row,
            'DOMICILIOS_TOTAL':        row[1] or 0,
            'DOM_SEM_CONJUGE_TOTAL':   row[13] or 0,
            'DOM_SEM_CONJUGE_HOMEM':   row[14] or 0,
            'DOM_SEM_CONJUGE_MULHER':  row[15] or 0,
        })

    print(f'{uf}: ok')

df_dom = pd.DataFrame(frames)

print(df_dom.shape)
print(df_dom.head())

df_dom.to_csv('domicilios_nordeste_2022.csv', sep=';', encoding='utf-8', index=False)