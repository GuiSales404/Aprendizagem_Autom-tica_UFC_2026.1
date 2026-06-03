from openpyxl import load_workbook
import pandas as pd

NORDESTE_CODIGOS = ['21', '22', '23', '24', '25', '26', '27', '28', '29']

wb = load_workbook('rendimentos.xlsx', read_only=True)
ws = wb['Tabela']

rows = []
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i < 5:
        continue

    cod  = str(row[0]) if row[0] else ''
    nome = str(row[1]) if row[1] else ''

    # Filtra só municípios — código de 7 dígitos e tem parênteses no nome
    if len(cod) != 7 or '(' not in nome:
        continue

    # Filtra Nordeste
    if cod[:2] not in NORDESTE_CODIGOS:
        continue

    uf   = nome.split('(')[-1].replace(')', '').strip()
    nome = nome.split('(')[0].strip()

    rows.append({
        'CO_MUNICIPIO':              int(cod),
        'PERC_RENDA_TRABALHO':       row[3] or 0,
        'PERC_RENDA_OUTRAS_FONTES':  row[4] or 0,
    })

df_renda_comp = pd.DataFrame(rows)

print(df_renda_comp.shape)
print(df_renda_comp.head())

df_renda_comp.to_csv('composicao_renda_nordeste_2022.csv', sep=';', encoding='utf-8', index=False)