from openpyxl import load_workbook
import pandas as pd

NORDESTE = ['AL', 'BA', 'CE', 'MA', 'PB', 'PE', 'PI', 'RN', 'SE']

wb = load_workbook('Tabela_5_Situacao_de_ocupacao.xlsx', read_only=True)
ws = wb['Tabela base do SIDRA 10253']

rows = []
for row in ws.iter_rows(values_only=True):
    val = row[0]
    if not val or '(' not in str(val):
        continue

    # Extrai UF do formato "Água Branca (AL)"
    uf = str(val).split('(')[-1].replace(')', '').strip()
    if uf not in NORDESTE:
        continue

    nome = str(val).split('(')[0].strip()

    rows.append({
        'NO_MUNICIPIO_IBGE': nome,
        'SG_UF': uf,
        'TOTAL_HOMENS':     row[4],
        'HOMENS_OCUPADOS':  row[5],
        'TOTAL_MULHERES':   row[7],
        'MULHERES_OCUPADAS': row[8],
    })

df_ocup = pd.DataFrame(rows)

print(df_ocup.shape)
print(df_ocup.head())

df_ocup.to_csv('ocupacao_nordeste_2022.csv', sep=';', encoding='utf-8', index=False)